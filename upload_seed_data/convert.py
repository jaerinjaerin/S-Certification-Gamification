import os
import msoffcrypto
from openpyxl import load_workbook
import pandas as pd
import json
from io import BytesIO

# 입력 및 출력 디렉토리 설정
input_dir = "data/origins/v5"
output_dir = "data/questions/update_new"

# 출력 디렉토리가 없으면 생성
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# 파일이 암호화되어 있는지 확인하는 함수
def is_encrypted(file_path):
    try:
        with open(file_path, 'rb') as f:
            office_file = msoffcrypto.OfficeFile(f)
            return office_file.is_encrypted()
    except Exception as e:
        print(f"Error checking encryption for {file_path}: {e}")
        return False

# 암호화된 파일을 해제하는 함수
def decrypt_excel(file_path, password):
    decrypted = BytesIO()
    with open(file_path, 'rb') as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
    return decrypted

# 파일 처리 함수 정의
def process_excel(file_path, output_path, password=None):
    # 파일 암호화 여부 체크 및 처리
    if is_encrypted(file_path):
        if password:
            decrypted_file = decrypt_excel(file_path, password)
            workbook = load_workbook(decrypted_file)
        else:
            print(f"File {file_path} is encrypted but no password was provided.")
            return
    else:
        workbook = load_workbook(file_path)

    sheet = workbook.active

    # 병합된 셀 정보 가져오기
    merged_cells = sheet.merged_cells.ranges
    merged_data = {}
    for merge_range in merged_cells:
        top_left_cell = merge_range.start_cell
        value = sheet[top_left_cell.coordinate].value
        for row in sheet.iter_rows(min_row=merge_range.min_row, max_row=merge_range.max_row,
                                   min_col=merge_range.min_col, max_col=merge_range.max_col):
            for cell in row:
                merged_data[cell.coordinate] = value

    # 데이터프레임으로 변환
    data = []
    found_no_column = False  # "No" 또는 "No"가 나올 때까지 행을 스킵하기 위한 플래그

    for row in sheet.iter_rows():  # values_only=True 제거
        # 병합된 값 적용
        row_data = [merged_data.get(cell.coordinate, cell.value) for cell in row]

        # "No" 또는 "No"가 포함된 행이 발견되면 플래그를 True로 설정
        if not found_no_column:
            if any(cell_value in ["No"] for cell_value in row_data):
                found_no_column = True
            else:
                continue  # 플래그가 False면 해당 행 스킵

        # 비어 있는 행은 추가하지 않음
        if all(cell_value is None for cell_value in row_data):
            continue

        data.append(row_data)

    # 비어 있는 파일 방지
    if not data:
        print(f"File {file_path} is empty after processing.")
        return

    # 데이터프레임 생성
    df = pd.DataFrame(data)

    # 비어 있는 열 확인 및 제거
    df = df.dropna(axis=1, how="all")

    # 비어 있는 데이터프레임 방지
    if df.empty:
        print(f"File {file_path} is empty after removing blank rows and columns.")
        return

    try:
        # 첫 번째 행을 헤더로 설정
        df.columns = df.iloc[0]
        df = df[1:].reset_index(drop=True)

        # 특정 열의 머리말이 없으면 기본값 설정
        if None in df.columns:
            df.columns = [
                "Question" if col is None and i == 10 else  # K열의 인덱스는 10 (0부터 시작)
                "Answer" if col is None and i == 11 else  # L열의 인덱스는 11
                col
                for i, col in enumerate(df.columns)
            ]

    except Exception as e:
        print(f"Error setting header for {file_path}: {e}")
        return

    # "No" 열이 숫자인지 확인 및 변환
    if "No" in df.columns:
        df["No"] = pd.to_numeric(df["No"], errors="coerce")
    # elif "No" in df.columns:
    #     df["No"] = pd.to_numeric(df["No"], errors="coerce")
    #     df = df.rename(columns={"No": "No"})  # "No"를 "No"로 변경
    # elif "No" in df.columns:
    #     df["No"] = pd.to_numeric(df["No"], errors="coerce")
    #     df = df.rename(columns={"No": "No"})  # "No"를 "No"로 변경
    else:
        print(f"Neither 'No' nor 'NO' column found in {file_path}. Skipping.")
        return

    # "No" 열 기준으로 그룹화 및 JSON 변환
    grouped_data = df.groupby("No").apply(lambda group: {
        "originQuestionIndex": int(group.name),  # 그룹화 기준 키 값을 사용
        "orderInStage": group["New No"].iloc[0] if "New No" in group else None,
        "enabled": group["Enabled"].iloc[0] if "Enabled" in group else None,
        "stage": group["Stage"].iloc[0] if "Stage" in group else None,
        "product": group["Product"].iloc[0] if "Product" in group else None,
        "category": group["Category"].iloc[0] if "Category" in group else None,
        "specificFeature": group["SpecificFeature"].iloc[0] if "SpecificFeature" in group else None,
        "importance": group["Importance"].iloc[0] if "Importance" in group else None,
        "timeLimitSeconds": group["TimeLimitSeconds"].iloc[0] if "TimeLimitSeconds" in group else None,
        "text": group["Question"].iloc[0] if "Question" in group else None,
        "questionType": group["QuestionType"].iloc[0] if "QuestionType" in group else None,
        "options": [
            {
                "text": row["Answer"] if "Answer" in row else None,
                "answerStatus": row["AnswerStatus"] if "AnswerStatus" in row else None,
            }
            for _, row in group.iterrows()
        ]
    }).tolist()

    # JSON 파일로 저장
    with open(output_path, 'w', encoding='utf-8') as f:
        try:
            # JSON 데이터를 문자열로 변환
            json_string = json.dumps(grouped_data, ensure_ascii=False, indent=2)
            # 파일에 저장
            f.write(json_string)
        except Exception as e:
            logging.error(f"Error saving JSON for file {file_path}: {e}")
            print(f"Error saving JSON for file {file_path}. Check error.log for details.")

# origins 디렉토리의 모든 엑셀 파일 처리
password = "Para2025!"  # 암호를 설정하거나 입력받는 방식으로 변경 가능
for file_name in os.listdir(input_dir):
    if file_name.endswith(".xlsx"):
        # 파일 이름에서 {와 } 제거
        sanitized_file_name = file_name.replace("{", "").replace("}", "")
        
        # 파일명 처리: "sample.NAT_021502.fi.xlsx" -> "NAT_021502.fi.json"
        parts = sanitized_file_name.split(".")
        if len(parts) > 2:  # 파일명이 "sample.NAT_021502.fi.xlsx" 구조라고 가정
            output_file_name = ".".join(parts[1:]).replace(".xlsx", ".json")
            input_path = os.path.join(input_dir, file_name)
            output_path = os.path.join(output_dir, output_file_name)

            try:
                # 엑셀 파일 처리
                process_excel(input_path, output_path, password)
                print(f"Processed: {sanitized_file_name} -> {output_file_name}")
            except Exception as e:
                print(f"Error processing {file_name}: {e}")

print("모든 파일이 변환되었습니다.")
